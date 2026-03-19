import streamlit as st
import openpyxl
from openpyxl.formula.translate import Translator
from openpyxl.utils.cell import range_boundaries, get_column_letter
from openpyxl.styles import Border, Side
from openpyxl.formatting.rule import Rule, IconSet, FormatObject
import io
import zipfile

def get_students_from_sheet(sheet):
    students = []
    start_reading = False
    for row in sheet.iter_rows(values_only=True):
        if row[1] == "STUDENT NUMBER":
            start_reading = True
            continue
        if start_reading:
            if not row[0] or not str(row[0]).strip().isdigit():
                break
            students.append({
                "index": row[0],
                "number": row[1],
                "name": row[2],
                "surname": row[3]
            })
    return students

def adjust_template_rows_and_tables(ws, num_students):
    start_row = 3
    current_rows = 30 # Varsayılan güvenlik değeri
    
    # 1. Şablonun orijinal tablosundaki MEVCUT satır sayısını dinamik olarak bul
    for table in ws.tables.values():
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        if min_row <= start_row <= max_row:
            current_rows = max_row - start_row + 1
            break
            
    # İşlemin yapılacağı orta noktayı dinamik belirle
    action_row_idx = start_row + (current_rows // 2)
    if action_row_idx <= start_row:
        action_row_idx = start_row + 1
    
    if num_students > current_rows:
        rows_to_add = num_students - current_rows
        ws.insert_rows(action_row_idx, amount=rows_to_add)
        
        for r in range(action_row_idx, action_row_idx + rows_to_add):
            for col in range(1, ws.max_column + 1):
                source_cell = ws.cell(row=action_row_idx - 1, column=col)
                target_cell = ws.cell(row=r, column=col)
                if source_cell.has_style:
                    target_cell._style = source_cell._style

    elif num_students < current_rows:
        rows_to_delete = current_rows - num_students
        ws.delete_rows(action_row_idx, amount=rows_to_delete)

    last_student_row = start_row + num_students - 1

    for table in ws.tables.values():
        ref = table.ref
        min_col, min_row, max_col, max_row = range_boundaries(ref)
        original_data_end = start_row + current_rows - 1
        offset = max_row - original_data_end
        if offset < 0:
            offset = 0
        new_table_max_row = last_student_row + offset
        table.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_table_max_row}"

    for r in range(start_row + 1, last_student_row + 1):
        for col in range(1, ws.max_column + 1):
            master_cell = ws.cell(row=start_row, column=col)
            target_cell = ws.cell(row=r, column=col)
            
            # SADECE formülleri taşı, statik rakamları kopyalama
            if master_cell.data_type == 'f' and master_cell.value:
                try:
                    target_cell.value = Translator(master_cell.value, origin=master_cell.coordinate).translate_formula(target_cell.coordinate)
                except:
                    target_cell.value = master_cell.value

    # E Sütunu (5. Sütun) Çerçeve Formatlaması
    thin_side = Side(border_style="thin")
    thick_side = Side(border_style="medium")
    
    for r in range(start_row, last_student_row + 1):
        cell = ws.cell(row=r, column=5)
        cell.border = Border(
            top=thick_side if r == start_row else thin_side,
            bottom=thick_side if r == last_student_row else thin_side,
            left=thick_side,
            right=thick_side
        )

    # --- YENİ KOŞULLU BİÇİMLENDİRME: 5'Lİ OK SİSTEMİ (Sıfırdan Enjekte) ---
    # 1. Eski sorunlu kuralları tamamen temizle
    if hasattr(ws.conditional_formatting, '_cf_rules'):
        ws.conditional_formatting._cf_rules.clear()
        
    # 2. Resimdeki barajlara göre (32, 24, 16, 8, 0) ok kuralını oluştur
    cfvo1 = FormatObject(type='num', val=0)   # Kırmızı Aşağı Ok (< 8)
    cfvo2 = FormatObject(type='num', val=8)   # Turuncu Sağ-Aşağı Ok (>= 8)
    cfvo3 = FormatObject(type='num', val=16)  # Sarı Sağ Ok (>= 16)
    cfvo4 = FormatObject(type='num', val=24)  # Sarı-Yeşil Sağ-Yukarı Ok (>= 24)
    cfvo5 = FormatObject(type='num', val=32)  # Yeşil Yukarı Ok (>= 32)
    
    icon_set = IconSet(iconSet='5Arrows', cfvo=[cfvo1, cfvo2, cfvo3, cfvo4, cfvo5])
    rule = Rule(type='iconSet', iconSet=icon_set)
    
    # 3. Kuralı sadece hedeflenen hücrelere uygula (E3'ten son öğrenciye kadar)
    ws.conditional_formatting.add(f"E3:E{last_student_row}", rule)

def process_class_template(template_bytes, class_name, students):
    wb = openpyxl.load_workbook(filename=io.BytesIO(template_bytes))
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        if ws["A1"].value:
            ws["A1"] = f"{class_name} {str(ws['A1'].value).split(' ', 1)[-1] if ' ' in str(ws['A1'].value) else ''}"
        
        adjust_template_rows_and_tables(ws, len(students))
        
        if sheet_name == "MidTerm":
            start_row = 3
            for i, student in enumerate(students):
                ws.cell(row=start_row + i, column=1, value=student["index"])
                ws.cell(row=start_row + i, column=2, value=student["number"])
                ws.cell(row=start_row + i, column=3, value=student["name"])
                ws.cell(row=start_row + i, column=4, value=student["surname"])
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()

st.title("Excel Grading Workbook Generator")

class_lists_file = st.file_uploader("Class Lists (Excel)", type=["xlsx"])

col1, col2 = st.columns(2)
with col1:
    st.subheader("1st Checker Templates")
    a1_1st = st.file_uploader("A1 1st Checker", type=["xlsx"])
    a2_1st = st.file_uploader("A2 1st Checker", type=["xlsx"])
    b1_1st = st.file_uploader("B1 1st Checker", type=["xlsx"])
    b2_1st = st.file_uploader("B2 1st Checker", type=["xlsx"])

with col2:
    st.subheader("2nd Checker Templates")
    a1_2nd = st.file_uploader("A1 2nd Checker", type=["xlsx"])
    a2_2nd = st.file_uploader("A2 2nd Checker", type=["xlsx"])
    b1_2nd = st.file_uploader("B1 2nd Checker", type=["xlsx"])
    b2_2nd = st.file_uploader("B2 2nd Checker", type=["xlsx"])

if st.button("Generate Workbooks"):
    templates = {
        "A1": {"1st": a1_1st, "2nd": a1_2nd},
        "A2": {"1st": a2_1st, "2nd": a2_2nd},
        "B1": {"1st": b1_1st, "2nd": b1_2nd},
        "B2": {"1st": b2_1st, "2nd": b2_2nd}
    }
    
    missing_files = False
    if not class_lists_file:
        missing_files = True
    
    if missing_files:
        st.error("Lütfen Class Lists dosyasını yükleyin.")
    else:
        with st.spinner("Dosyalar oluşturuluyor..."):
            class_wb = openpyxl.load_workbook(class_lists_file, data_only=True)
            
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                for sheet_name in class_wb.sheetnames:
                    level = sheet_name.split(".")[0]
                    
                    if level in templates:
                        ws = class_wb[sheet_name]
                        students = get_students_from_sheet(ws)
                        
                        if not students:
                            continue
                            
                        if templates[level]["1st"]:
                            file_1st = process_class_template(templates[level]["1st"].getvalue(), sheet_name, students)
                            zip_file.writestr(f"{level}/{sheet_name} 1st Checker.xlsx", file_1st)
                            
                        if templates[level]["2nd"]:
                            file_2nd = process_class_template(templates[level]["2nd"].getvalue(), sheet_name, students)
                            zip_file.writestr(f"{level}/{sheet_name} 2nd Checker.xlsx", file_2nd)

            zip_buffer.seek(0)
            st.success("Tüm dosyalar başarıyla oluşturuldu!")
            st.download_button(
                label="Oluşturulan Dosyaları İndir (ZIP)",
                data=zip_buffer,
                file_name="Grading_Workbooks.zip",
                mime="application/zip"
            )
